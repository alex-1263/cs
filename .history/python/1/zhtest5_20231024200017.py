import openpyxl
Workbook1 = openpyxl.Workbook() 
worksheet1 = Workbook1.create_sheet(title = "成绩" )
i=1
while(1):
    name=input("请输入考生姓名：")
    subject=input("请输入考试科目：")
    grade=input("请输入考试成绩：")
    worksheet1.cell(i,1,name)
    worksheet1.cell(i,2,subject)
    worksheet1.cell(i,3,grade)
    Workbook1.save(filename="grade.xlsx")
    i+=1
    if(i>3):
        break


