import openpyxl
workbook1=openpyxl.Workbook()
workbook=openpyxl.load_workbook("C:/Users/Alex1/Downloads/各学院汇报数据.xlsx")
shenames=workbook.sheetnames
wworksheet=workbook[shenames[0]]
wworksheet1=workbook.worksheets[0]
wworksheet2=workbook.active
name=wworksheet.title
rows=wworksheet.max_row
columns=wworksheet.max_column
