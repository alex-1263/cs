import openpyxl
Workbook1 = openpyxl.Workbook() 
worksheet1 = Workbook1.create_sheet(title = "成绩" )
worksheet2 = Workbook1.create_sheet(title = "期末" )
worksheet2.append(["姓名","科目","成绩"])
i=1
while(1):
    name=input("请输入考生姓名：")
    subject=input("请输入考试科目：")
    grade=input("请输入考试成绩：")
    worksheet1.cell(i,1,name)
    worksheet1.cell(i,2,subject)
    worksheet1.cell(i,3,grade)
    Workbook1.save(filename="grade.xlsx")
    i+=1
    flag=input("请输入end结束录入成绩：")
    if flag=='end':
        break
for row in worksheet1.rows:
    for cell in row:
        print(cell.value,end=" ")
    print("")
rows=worksheet1.max_row
name_set=set()
for i in range(1,rows):
    name_set.add(worksheet1.cell(i,1).value)
subject_set=set()
for i in len(name_set):
    for j in range(1,rows):
        if name_set[i]==worksheet1.cell(j,1).value:
            subject_set.add(worksheet1.cell(j,2).value)
    for j in len(subject_set):
        grade=0
        for k in range(1,rows): 
            if name_set[i]==worksheet1.cell(k,1).value and subject_set[j]==worksheet1.cell(k,2).value:
                if worksheet1.cell(k,3).value>grade:
                    grade=worksheet1.cell(k,3).value

            




